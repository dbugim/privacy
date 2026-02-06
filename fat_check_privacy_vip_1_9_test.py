# Standard library imports
import os
from openpyxl import load_workbook
import glob
import time
import subprocess
from datetime import datetime, timedelta

# Third-party imports
from playwright.sync_api import sync_playwright

# Global variable to store the result
privacy_vip_income = 0.0

def insert_username_and_password_and_enter(page):
    try:
        print("Waiting for email field...")
        email_input = page.locator('input[type="email"]')
        email_input.fill('hacksimone29@gmail.com')
        print("Email inserted")
        page.wait_for_timeout(2000)
    except:
        print("Could not insert email. Continuing...")
        return

    try:
        print("Waiting for password field...")
        password_input = page.locator('input[type="password"]')
        password_input.fill('#Partiu15')
        print("Password inserted")
        page.wait_for_timeout(2000)
    except:
        print("Could not insert password. Continuing...")
        return

    try:
        print("Clicking Enter button...")
        button = page.locator('button:has-text("Entrar")')
        for i in range(3):
            try:
                button.click()
                print(f"Click {i+1} performed")
                page.wait_for_timeout(2000)
                break
            except:
                print(f"Click {i+1} failed. Retrying...")
                page.wait_for_timeout(1000)
    except:
        print("Could not click Enter button. Continuing...")
        return

#def calculate_yesterday_income(page) -> float:
    """Calculates the sum of the 'Your commission' column for yesterday's date - navigating through pagination"""
    yesterday = datetime.now() - timedelta(days=1)
    yesterday_str = yesterday.strftime('%d/%m/%Y')
    yesterday_str_alt = yesterday.strftime('%d/%m/%y')   # alternative format

    total = 0.0
    found_rows = 0
    page_num = 1
    max_pages_without_yesterday = 10  # safety: stop if yesterday doesn't appear in many pages

    # Selectors for data rows (excluding header)
    row_selectors = [
        '#pane-statement .el-table__body-wrapper .el-table__body tr.el-table__row',
        '#pane-statement .el-table__body tr.el-table__row',
        '.el-table .el-table__body tr.el-table__row',
        '.el-table__row'
    ]

    # Selectors for "next" button in pagination (common in Element UI)
    next_button_selectors = [
        '.el-pagination .btn-next:not(.is-disabled)',
        '.el-pagination__jump-next',  # if there is ellipsis
        '.el-pagination button.btn-next'
    ]

    print("Starting table reading with pagination...")

    while True:
        # Find the working row selector
        rows = []
        row_sel = None
        for sel in row_selectors:
            rows = page.query_selector_all(sel)
            if len(rows) > 0:
                row_sel = sel
                print(f"→ Rows found on page {page_num} using: {sel} ({len(rows)} rows)")
                break

        if not rows:
            print("No rows found on the current page!")
            break

        # Process rows on the current page
        for row in rows:
            cells = row.query_selector_all('td')
            if len(cells) < 4:
                continue

            date_text = cells[0].inner_text().strip()
            if not date_text:
                continue

            if date_text not in (yesterday_str, yesterday_str_alt):
                continue

            # "Your commission" column - index 3
            commission_cell = cells[3]
            commission_text = commission_cell.inner_text().strip()

            # Robust cleaning for Brazilian format
            cleaned = (
                commission_text
                .replace('R$', '')
                .replace(' ', '')
                .replace('.', '')      # thousands
                .replace(',', '.')     # decimal
            )

            try:
                value = float(cleaned)
                total += value
                found_rows += 1
                print(f"  ✓ {date_text} (page {page_num}) → R$ {value:,.2f}")
            except ValueError:
                print(f"  Invalid value ignored: {commission_text!r}")

        # Check if found yesterday's rows; if not, and many pages, stop early
        if found_rows == 0:
            max_pages_without_yesterday -= 1
            if max_pages_without_yesterday <= 0:
                print("→ Too many pages without yesterday's data. Stopping to avoid excessive loop.")
                break

        # Find the "next" button
        next_button = None
        for sel in next_button_selectors:
            btn = page.query_selector(sel)
            if btn and btn.is_visible() and not btn.get_attribute('disabled'):
                next_button = btn
                print(f"→ Next button found: {sel}")
                break

        if not next_button:
            print(f"→ End of pagination on page {page_num}.")
            break

        # Click next and wait for loading
        next_button.click()
        time.sleep(1.5)  # initial pause for click

        # Wait until table rows are updated (new rows appear)
        try:
            page.wait_for_selector(row_sel, timeout=8000)  # wait up to 8s
            # Check if number of rows is ~20 to confirm load
            new_rows = page.query_selector_all(row_sel)
            while len(new_rows) < 15:  # tolerance if less than 20 on last
                time.sleep(0.8)
                new_rows = page.query_selector_all(row_sel)
        except Exception as e:
            print(f"→ Error waiting for new page: {e}. Stopping.")
            break

        page_num += 1
        print(f"→ Moving to page {page_num} | total so far: R$ {total:,.2f}")

    if found_rows == 0:
        print(f"No rows found for yesterday ({yesterday_str})")
        # Debug: show some dates from the first page
        page.go_back()  # go back to first if necessary, but assuming already processed
        rows = page.query_selector_all(row_selectors[0])
        print("Some dates found (first 8):")
        for row in rows[:8]:
            cells = row.query_selector_all('td')
            if cells:
                print("   ", cells[0].inner_text().strip())

    print(f"\nFinished. Found {found_rows} rows from yesterday (expected ~20). Total: R$ {total:,.2f}")
    return total

def launch_chrome_on_with_privacy_vip_login_page():
    chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    user_data_dir = os.path.join(os.environ['LOCALAPPDATA'], r"Google\Chrome\User Data\AutomationProfile")

    print("Terminating existing Chrome instances...")
    os.system("taskkill /f /im chrome.exe /t >nul 2>&1")
    time.sleep(3)

    if not os.path.exists(chrome_path):
        raise FileNotFoundError(f"Chrome not found at: {chrome_path}")

    print("Starting Chrome with remote debugging...")
    process = subprocess.Popen([
        chrome_path,
        f"--user-data-dir={user_data_dir}",
        "--remote-debugging-port=9222",
        "--start-maximized",
        "--no-first-run",
        "--no-default-browser-check",
        "https://privacy.com.br/"
    ])
    time.sleep(7)

    pw = sync_playwright().start()
    browser = pw.chromium.connect_over_cdp("http://localhost:9222")
    context = browser.contexts[0]
    page = context.pages[0] if context.pages else context.new_page()

    # Enhanced manual stealth measures (replacing playwright_stealth to avoid deprecation warnings)
    page.add_init_script("""
        // Basic anti-detection
        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
        Object.defineProperty(navigator, 'languages', { get: () => ['pt-BR', 'pt', 'en-US', 'en'] });
        Object.defineProperty(navigator, 'plugins', { get: () => [1,2,3,4,5] });
        Object.defineProperty(navigator, 'platform', { get: () => 'Win32' });
        Object.defineProperty(navigator, 'userAgent', { get: () => 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36' });
        Object.defineProperty(navigator, 'hardwareConcurrency', { get: () => 4 });
        Object.defineProperty(navigator, 'deviceMemory', { get: () => 8 });

        // Mock Chrome runtime
        window.chrome = {
            runtime: {},
            loadTimes: () => ({ requestTime: Date.now() / 1000 }),
            csi: () => ({ onloadT: Date.now(), pageT: performance.now() })
        };

        // Mock permissions
        const originalQuery = navigator.permissions.query;
        navigator.permissions.query = (parameters) => (
            parameters.name === 'notifications' ?
                Promise.resolve({ state: Notification.permission }) :
                originalQuery(parameters)
        );

        // Mock canvas fingerprinting
        const getContext = HTMLCanvasElement.prototype.getContext;
        HTMLCanvasElement.prototype.getContext = function(type) {
            const ctx = getContext.apply(this, arguments);
            if (type === '2d') {
                ctx.getImageData = function(x, y, w, h) {
                    const data = ctx.getImageData(x, y, w, h);
                    // Add noise to data to alter fingerprint
                    for (let i = 0; i < data.data.length; i += 4) {
                        data.data[i] += Math.random() * 2 - 1;
                        data.data[i+1] += Math.random() * 2 - 1;
                        data.data[i+2] += Math.random() * 2 - 1;
                    }
                    return data;
                };
            }
            return ctx;
        };

        // Mock WebGL
        const getParameter = WebGLRenderingContext.prototype.getParameter;
        WebGLRenderingContext.prototype.getParameter = function(parameter) {
            if (parameter === 37445) return 'Intel Inc.';
            if (parameter === 37446) return 'Intel Iris OpenGL Engine';
            return getParameter.apply(this, arguments);
        };
    """)

    print("Connection established with Chrome!")
    return pw, context, page, process

def try_close_popup(page):
    selectors = [
        'button:has-text("Fechar")',
        'button[aria-label*="fechar" i]',
        ".close-icon",
        "#privacy-web-stories >> button",
        'button:has(.fa-xmark)',
    ]
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.is_visible(timeout=2000):
                loc.click(timeout=5000)
                print("Popup closed:", sel)
                time.sleep(1)
                return True
        except:
            continue
    return False

def click_extrato_tab(page):
    selectors = [
        '#tab-statement',
        '.el-tabs__item#tab-statement',
        'div[aria-controls="pane-statement"]',
        '//*[contains(text(),"Extrato")]',
        'button:has-text("Extrato")',
        '#privacy-web-myprivacy >> #tab-statement',
    ]
    for selector in selectors:
        try:
            if selector.startswith('//'):
                loc = page.locator(f"xpath={selector}")
            else:
                loc = page.locator(selector)
            if loc.count() > 0 and loc.is_visible(timeout=4000):
                loc.first.click(timeout=8000)
                print(f"'Statement' tab clicked using: {selector}")
                return True
        except:
            continue
    print("Could not locate the Statement tab")
    return False

def cleanup(pw=None, context=None, browser_process=None):
    if context:
        try: context.close()
        except: pass
    if pw:
        try: pw.stop()
        except: pass
    if browser_process:
        try:
            browser_process.terminate()
            browser_process.wait(timeout=5)
        except: pass
    print("Resources released")

def click_on_calendar(page):
    """
    Attempt to find and click the Calendar icon using multiple approaches,
    specifically handling elements inside Shadow DOM.
    """
    try:
        # List of selectors to try
        # Note: Playwright's locator() can often pierce Shadow DOM automatically with CSS,
        # but we include the explicit JS Path for safety.
        selectors = [
            # Direct CSS selector (Playwright usually pierces shadow roots with this)
            "i.el-icon.el-input__icon.el-range__icon",

            # Specific Path CSS
            "#pane-statement i.el-range__icon",

            # XPath
            "//*[@id='pane-statement']//i[contains(@class, 'el-range__icon')]",

            # The full JS Path provided (Direct Shadow DOM access)
            'document.querySelector("#privacy-web-myprivacy").shadowRoot.querySelector("#pane-statement > div > div:nth-child(1) > div > div.card-body > div.border-0 > div > div > div:nth-child(1) > div > i.el-icon.el-input__icon.el-range__icon")'
        ]

        for selector in selectors:
            try:
                # Handle the explicit Shadow Root JS selector
                if selector.startswith("document.querySelector"):
                    button_clicked = page.evaluate(f'''() => {{
                        const element = {selector};
                        if (element) {{
                            element.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            element.click();
                            return true;
                        }}
                        return false;
                    }}''')
                    if button_clicked: return True

                # Handle XPath
                elif selector.startswith('//') or selector.startswith('(*'):
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        xpath_elements.first.click(force=True)
                        return True

                # Handle Standard CSS (Playwright auto-pierces Shadow DOM)
                else:
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        css_elements.first.click(force=True)
                        return True

            except Exception:
                continue

        # Fallback JavaScript approach specifically for Element UI / Shadow DOM
        fallback_clicked = page.evaluate('''() => {
            // Helper to find element inside shadow roots recursively
            const findInShadow = (selector) => {
                let result = null;
                const search = (root) => {
                    if (root.querySelector(selector)) {
                        result = root.querySelector(selector);
                        return;
                    }
                    const shadows = Array.from(root.querySelectorAll('*')).filter(el => el.shadowRoot);
                    for (let s of shadows) {
                        search(s.shadowRoot);
                        if (result) return;
                    }
                };
                search(document);
                return result;
            };

            const calendarIcon = findInShadow('i.el-range__icon') || findInShadow('.el-icon-date');
            if (calendarIcon) {
                calendarIcon.scrollIntoView({behavior: 'auto', block: 'center'});
                calendarIcon.click();
                return true;
            }
            return false;
        }''')

        return fallback_clicked

    except Exception as e:
        print(f"Error in click_on_calendar: {str(e)}")
        return False
def select_last_seven_days_range(page):
    """
    Calculates the date 7 days ago (oldest) and yesterday, then clicks them in the Element calendar to select the range.
    Handles Shadow DOM and cross-month date selection.
    """
    try:
        # 1. Calculate dates
        yesterday = datetime.now() - timedelta(days=1)
        seven_days_ago = datetime.now() - timedelta(days=7)

        # 2. JavaScript to click both dates (including prev-month cells if needed)
        click_script = f'''() => {{
            const findInShadow = (selector, root = document) => {{
                const el = root.querySelector(selector);
                if (el) return el;
                const shadows = Array.from(root.querySelectorAll('*')).filter(e => e.shadowRoot);
                for (let s of shadows) {{
                    const result = findInShadow(selector, s.shadowRoot);
                    if (result) return result;
                }}
                return null;
            }};

            // Find ALL available date cells (including prev-month and next-month)
            const cells = Array.from(document.querySelectorAll('.el-date-table td.available'))
                .concat(Array.from(findInShadow('.el-date-table') ?
                    findInShadow('.el-date-table').querySelectorAll('td.available') : []));

            // Match by day number - find the start date (older, should appear first in DOM if both months visible)
            const targetStart = cells.find(cell => {{
                const text = cell.innerText.trim();
                return text === "{seven_days_ago.day}";
            }});

            // Match by day number - find the end date (yesterday)
            const targetEnd = cells.reverse().find(cell => {{
                const text = cell.innerText.trim();
                return text === "{yesterday.day}";
            }});

            if (targetStart && targetEnd) {{
                targetStart.scrollIntoView({{behavior: 'auto', block: 'center'}});
                targetStart.click();
                setTimeout(() => {{
                    targetEnd.scrollIntoView({{behavior: 'auto', block: 'center'}});
                    targetEnd.click();
                }}, 200);
                return true;
            }}
            return false;
        }}'''

        success = page.evaluate(click_script)
        return success
    except Exception as e:
        print(f"Error in select_last_seven_days_range: {str(e)}")
        return False


def click_on_extrato_de_venda_next_page_button(page):
    """
    Finds and clicks the 'Next Page' button in the sales statement pagination.
    Handles Shadow DOM, checks for disabled states, and waits for UI transition.
    """
    try:
        # 1. Advanced JavaScript Evaluation for Shadow DOM and State
        # This is the most reliable method for Element UI components inside Shadow Roots
        js_click_script = '''() => {
            const findElementInShadows = (selector, root = document) => {
                const el = root.querySelector(selector);
                if (el) return el;
                const shadows = Array.from(root.querySelectorAll('*')).filter(e => e.shadowRoot);
                for (let s of shadows) {
                    const found = findElementInShadows(selector, s.shadowRoot);
                    if (found) return found;
                }
                return null;
            };

            const btn = findElementInShadows('button.btn-next');

            if (!btn) return "not_found";

            // Check if button is disabled via attribute, property, or CSS class
            const isReadonly = btn.disabled ||
                               btn.getAttribute('aria-disabled') === 'true' ||
                               btn.classList.contains('disabled');

            if (isReadonly) return "disabled";

            btn.scrollIntoView({behavior: 'auto', block: 'center'});
            btn.click();
            return "clicked";
        }'''

        result = page.evaluate(js_click_script)

        if result == "clicked":
            # Mandatory wait for the table to begin refreshing
            page.wait_for_timeout(2000)
            return True
        elif result == "disabled":
            print("Pagination: Reached the last page (Next button is disabled).")
            return False
        else:
            # Fallback to Playwright's native locators if JS didn't find it
            # (Playwright pierces Shadow DOM automatically with CSS selectors)
            native_btn = page.locator("#pane-statement button.btn-next").first
            if native_btn.is_visible() and native_btn.is_enabled():
                native_btn.click(force=True)
                page.wait_for_timeout(2000)
                return True

            print("Pagination: Next page button not found.")
            return False

    except Exception as e:
        print(f"Error in click_on_extrato_de_venda_next_page_button: {str(e)}")
        return False

def click_on_gerar_relatorio_button(page):
    """
    Finds and clicks the 'Generate Report' button in the sales statement.
    Handles Shadow DOM and checks for disabled state.
    """
    try:
        # List of selectors specific to the Generate Report button
        selectors = [
            # 1. Direct CSS (Playwright automatically pierces Shadow DOM for CSS)
            "#pane-statement button.btn-primary:has-text('Gerar Relatório')",

            # 2. Your provided CSS Selector
            "#pane-statement > div > div:nth-child(1) > div > div.card-buttons > button",

            # 3. Text-based locator
            "button:has-text('Gerar Relatório')",

            # 4. Provided XPath
            "xpath=//*[@id='pane-statement']/div/div[1]/div/div[3]/button",

            # 5. Provided JS Path for Shadow DOM
            'document.querySelector("#privacy-web-myprivacy").shadowRoot.querySelector("#pane-statement > div > div:nth-child(1) > div > div.card-buttons > button")'
        ]

        for selector in selectors:
            try:
                # Handle JS Path specifically
                if selector.startswith("document.querySelector"):
                    clicked = page.evaluate(f'''() => {{
                        const btn = {selector};
                        if (btn && btn.getAttribute('aria-disabled') !== 'true') {{
                            btn.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            btn.click();
                            return true;
                        }}
                        return false;
                    }}''')
                    if clicked: return True

                # Handle Standard Locators (CSS/XPath)
                else:
                    loc = page.locator(selector).first
                    if loc.count() > 0 and loc.is_visible():
                        # Force visibility and click
                        loc.scroll_into_view_if_needed()
                        loc.click(force=True)
                        return True
            except:
                continue

        # Global Shadow DOM Fallback (Search for button by text inside all shadow roots)
        fallback_clicked = page.evaluate('''() => {
            const findInShadow = (root = document) => {
                const buttons = Array.from(root.querySelectorAll('button'));
                const target = buttons.find(b => b.innerText.includes('Gerar Relatório'));
                if (target && target.getAttribute('aria-disabled') !== 'true') return target;

                const shadows = Array.from(root.querySelectorAll('*')).filter(e => e.shadowRoot);
                for (let s of shadows) {
                    const found = findInShadow(s.shadowRoot);
                    if (found) return found;
                }
                return null;
            };
            const btn = findInShadow();
            if (btn) {
                btn.click();
                return true;
            }
            return false;
        }''')

        return fallback_clicked

    except Exception as e:
        print(f"Error in click_on_gerar_relatorio_button: {str(e)}")
        return False

def click_on_confirmar_button(page):
    """
    Finds and clicks the 'Confirm' button inside the 'Generate Report' overlay dialog.
    Handles the el-overlay-dialog structure specifically.
    """
    try:
        # 1. First, try to locate the specific dialog by its aria-label
        # This is the most reliable way to find this specific popup
        js_click_script = '''() => {
            const findButtonInDialog = (root = document) => {
                // 1. Find the dialog container by aria-label
                const dialog = root.querySelector('div[role="dialog"][aria-label="Gerar Relatório"]');

                if (dialog) {
                    // 2. Find the Confirm button within that dialog's footer
                    const buttons = Array.from(dialog.querySelectorAll('button.btn-primary'));
                    const confirmBtn = buttons.find(b => b.innerText.includes('Confirmar'));

                    if (confirmBtn && confirmBtn.getAttribute('aria-disabled') !== 'true') {
                        confirmBtn.click();
                        return "clicked";
                    }
                    return confirmBtn ? "disabled" : "button_not_found";
                }

                // 2. If not found, recurse into shadow roots
                const shadows = Array.from(root.querySelectorAll('*')).filter(e => e.shadowRoot);
                for (let s of shadows) {
                    const result = findButtonInDialog(s.shadowRoot);
                    if (result !== "not_found") return result;
                }
                return "not_found";
            };

            return findButtonInDialog();
        }'''

        result = page.evaluate(js_click_script)

        if result == "clicked":
            return True
        elif result == "disabled":
            print("Confirm button is disabled (check if form fields are filled).")
            return False

        # 3. Native Playwright Fallback (If JS fails to find the dialog)
        # Playwright's locator('role=dialog') is very powerful
        try:
            confirm_loc = page.get_by_role("dialog", name="Gerar Relatório").get_by_role("button", name="Confirmar")
            if confirm_loc.is_visible():
                confirm_loc.click(force=True)
                return True
        except:
            pass

        print("Could not find the 'Generate Report' confirmation dialog.")
        return False

    except Exception as e:
        print(f"Error in click_on_confirmar_button: {str(e)}")
        return False

def read_report():
    """
    Reads the most recent XLSX report, sums Column D (4th column),
    and updates the global privacy_vip_income variable.
    """
    global privacy_vip_income
    target_folder = r"C:\Users\danie\Desktop\Trampo\Faturamento\rels"

    try:
        # 1. Find the most recent .xlsx file
        list_of_files = glob.glob(os.path.join(target_folder, "*.xlsx"))
        if not list_of_files:
            print("No XLSX files found in the directory.")
            return False

        latest_file = max(list_of_files, key=os.path.getctime)
        print(f"Reading latest Excel report: {os.path.basename(latest_file)}")

        # 2. Load the Workbook
        # data_only=True ensures we get the calculated value, not the formula
        wb = load_workbook(latest_file, data_only=True)
        sheet = wb.active  # Gets the first visible sheet

        total_sum = 0.0

        # 3. Iterate through rows (starting from row 2 to skip header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Column D is index 3 in the resulting tuple
            if len(row) >= 4:
                value_raw = row[3]

                if value_raw is None:
                    continue

                try:
                    # openpyxl usually returns numbers as floats/ints automatically
                    # If it's a string (e.g., "R$ 1.200,50"), we clean it
                    if isinstance(value_raw, str):
                        clean_val = value_raw.replace('R$', '').replace('$', '').strip()
                        if ',' in clean_val and '.' in clean_val:
                            clean_val = clean_val.replace('.', '').replace(',', '.')
                        elif ',' in clean_val:
                            clean_val = clean_val.replace(',', '.')
                        total_sum += float(clean_val)
                    else:
                        total_sum += float(value_raw)

                except (ValueError, TypeError):
                    # Skip if the cell contains non-numeric text
                    continue

        # 4. Update the global variable
        privacy_vip_income = total_sum
        wb.close() # Close the file handle

        print(f"Calculation complete! Total Income: {privacy_vip_income:.2f}")
        return True

    except Exception as e:
        print(f"Error reading the Excel report: {str(e)}")
        return False

privacy_vip_income = 0.0

def main():
    global privacy_vip_income
    pw = None
    context = None
    browser_process = None
    pw, context, page, browser_process = launch_chrome_on_with_privacy_vip_login_page()

    # region Try to log in if not already logged in
    success = insert_username_and_password_and_enter(page)
    if success:
        print("Login completed!")
    else:
        print("Login attempt failed or already logged in.")

    # Proceed to navigate anyway
    page.wait_for_load_state("networkidle", timeout=15000)
    page.wait_for_timeout(5000)
    page.goto("https://privacy.com.br/myprivacy")
    page.wait_for_load_state("networkidle", timeout=15000)
    page.wait_for_timeout(3000)
    # endregion

    # region Close pop-ups with selector wait
    print("Checking/closing pop-ups...")
    for _ in range(5):
        if try_close_popup(page):
            page.wait_for_timeout(1200)
        else:
            break
    # endregion

    # region Open Statement tab with retries and wait
    print("\nTrying to open Statement tab...")
    extrato_success = False
    for attempt in range(1, 7):
        print(f"Attempt {attempt}/6...")
        if click_extrato_tab(page):
            extrato_success = True
            break
        time.sleep(2)
    if not extrato_success:
        print("!!! CRITICAL FAILURE: Could not open the Statement tab")
        page.screenshot(path="error_statement_tab.png")  # Screenshot for debugging
        return

    print("\nWaiting for complete loading of the statement table...")
    page.wait_for_selector('#pane-statement', state='visible', timeout=30000)
    time.sleep(3)  # Extra safety
    # endregion

    # region Try to click the Calendar icon
    max_retries = 3
    calendar_success = False

    for attempt in range(max_retries):
        if click_on_calendar(page):
            calendar_success = True
            break
        else:
            print(f"Calendar click attempt {attempt + 1} failed.")
            if attempt < max_retries - 1:
                time.sleep(1)

    if not calendar_success:
        print("Failed to click Calendar after all attempts.")
        page.screenshot(path="error_calendar.png")
    
    page.wait_for_timeout(20000)
    # endregion

    # region Try to select Last Seven Days Range with retries
    max_retries = 3
    for attempt in range(max_retries):
        if select_last_seven_days_range(page):
            print("Successfully selected Last Seven Days Range (from 7 days ago to yesterday).")
            break
        else:
            print(f"Attempt {attempt + 1} to select Last Seven Days Range failed.")
            time.sleep(1)
    else:
        print("Failed to select Last Seven Days Range after all attempts.")
        page.screenshot(path="error_last_seven_days.png")
    page.wait_for_timeout(40000)
    # endregion

    # region Try to click the Generate Report button
    if click_on_gerar_relatorio_button(page):
        print("Dialog 'Generate Report' opened.")
        time.sleep(2) # Wait for dialog animation

        # DO NOT call click_on_confirmar_button here!
        # It will be called inside the download listener below.
    # endregion

    # region Confirm and Download Report
    target_folder = r"C:\Users\danie\Desktop\Trampo\Faturamento\rels"
    import os
    os.makedirs(target_folder, exist_ok=True)

    max_retries = 3
    download_success = False

    for attempt in range(max_retries):
        try:
            print(f"Attempt {attempt + 1}: Waiting for download event...")

            # 1. Start the listener
            with page.expect_download(timeout=60000) as download_info:
                # 2. Trigger the click ONLY HERE
                if click_on_confirmar_button(page):
                    print("Confirm button clicked. Processing file...")

                    # 3. Capture and save the file
                    download = download_info.value
                    final_destination = os.path.join(target_folder, download.suggested_filename)
                    download.save_as(final_destination)

                    print(f"File successfully saved to: {final_destination}")
                    download_success = True
                    break
                else:
                    print(f"Click logic could not find the button on attempt {attempt + 1}")

        except Exception as e:
            print(f"Attempt {attempt + 1} failed: {str(e)}")
            page.screenshot(path=f"error_download_attempt_{attempt+1}.png")
            if attempt < max_retries - 1:
                time.sleep(2)

    if not download_success:
        print("Failed to capture the download after all retries.")
    # endregion

    time.sleep(5)
    # Read the downloaded report and calculate income
    read_report()

    print("Browser kept open. Press Ctrl+C to terminate.\n")
    while True:
        time.sleep(1.5)

if __name__ == "__main__":
    main()
