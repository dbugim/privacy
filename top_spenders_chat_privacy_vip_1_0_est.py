# Standard library imports
import os
import sys
from openpyxl import load_workbook, Workbook
import glob
import time
import subprocess
from datetime import datetime, timedelta

# Third-party imports
from playwright.sync_api import sync_playwright

# region playwright-stealth (fork mais atualizado recomendado em 2025/2026)
try:
    from playwright_stealth import stealth_sync
except ImportError:
    print("playwright-stealth não encontrado.")
    print("Instale com: pip install git+https://github.com/AtuboDad/playwright_stealth.git")
    sys.exit(1)
# endregion

def launch_chrome_with_debugging():
    chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    user_data_dir = os.path.join(os.environ['LOCALAPPDATA'], r"Google\Chrome\User Data\AutomationProfile")

    print("Terminating existing Chrome instances...")
    os.system("taskkill /f /im chrome.exe /t >nul 2>&1")
    time.sleep(3)

    if not os.path.exists(chrome_path):
        raise FileNotFoundError(f"Chrome not found at: {chrome_path}")

    print("Starting Chrome in HEADLESS mode...")
    # Headless mode with remote debugging enabled
    process = subprocess.Popen([
        chrome_path,
        f"--user-data-dir={user_data_dir}",
        "--remote-debugging-port=9222",
        "--headless=new",             # Enables the modern headless engine
        "--disable-gpu",              # Necessary for stability in headless environments
        "--window-size=1920,1080",    # Ensures the layout renders as a desktop screen
        "--no-first-run",
        "--no-default-browser-check",
        "https://privacy.com.br/myprivacy"
    ])
    time.sleep(7)

    pw = sync_playwright().start()
    browser = pw.chromium.connect_over_cdp("http://localhost:9222")
    context = browser.contexts[0]
    page = context.pages[0] if context.pages else context.new_page()
    stealth_sync(page)

    # Extra anti-detection measures for Headless mode
    page.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
        window.chrome = { runtime: {} };
        Object.defineProperty(navigator, 'languages', { get: () => ['en-US', 'en', 'pt-BR', 'pt'] });
        Object.defineProperty(navigator, 'plugins', { get: () => [1,2,3,4,5] });
        Object.defineProperty(navigator, 'platform', { get: () => 'Win32' });
    """)

    print("Connection established with Headless Chrome!")
    return pw, context, page, process

def try_close_popup(page):
    selectors = [
        'button:has-text("Fechar")',
        'button[aria-label*="fechar" i]',
        ".close-icon",
        "#privacy-web-stories >> button",
        'button:has(.fa-xmark)',
    ]
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.is_visible(timeout=2000):
                loc.click(timeout=5000)
                print("Popup fechado:", sel)
                time.sleep(1)
                return True
        except:
            continue
    return False

def click_extrato_tab(page):
    selectors = [
        '#tab-statement',
        '.el-tabs__item#tab-statement',
        'div[aria-controls="pane-statement"]',
        '//*[contains(text(),"Extrato")]',
        'button:has-text("Extrato")',
        '#privacy-web-myprivacy >> #tab-statement',
    ]
    for selector in selectors:
        try:
            if selector.startswith('//'):
                loc = page.locator(f"xpath={selector}")
            else:
                loc = page.locator(selector)
            if loc.count() > 0 and loc.is_visible(timeout=4000):
                loc.first.click(timeout=8000)
                print(f"Aba 'Extrato' clicada usando: {selector}")
                return True
        except:
            continue
    print("Não conseguiu localizar a aba Extrato")
    return False

def cleanup(pw=None, context=None, browser_process=None):
    if context:
        try: context.close()
        except: pass
    if pw:
        try: pw.stop()
        except: pass
    if browser_process:
        try:
            browser_process.terminate()
            browser_process.wait(timeout=5)
        except: pass
    print("Recursos liberados")

def click_on_calendar(page):
    """
    Attempt to find and click the Calendar icon using multiple approaches,
    specifically handling elements inside Shadow DOM.
    """
    try:
        # List of selectors to try
        # Note: Playwright's locator() can often pierce Shadow DOM automatically with CSS,
        # but we include the explicit JS Path for safety.
        selectors = [
            # Direct CSS selector (Playwright usually pierces shadow roots with this)
            "i.el-icon.el-input__icon.el-range__icon",
            
            # Specific Path CSS
            "#pane-statement i.el-range__icon",
            
            # XPath
            "//*[@id='pane-statement']//i[contains(@class, 'el-range__icon')]",
            
            # The full JS Path provided (Direct Shadow DOM access)
            'document.querySelector("#privacy-web-myprivacy").shadowRoot.querySelector("#pane-statement > div > div:nth-child(1) > div > div.card-body > div.border-0 > div > div > div:nth-child(1) > div > i.el-icon.el-input__icon.el-range__icon")'
        ]

        for selector in selectors:
            try:
                # Handle the explicit Shadow Root JS selector
                if selector.startswith("document.querySelector"):
                    button_clicked = page.evaluate(f'''() => {{
                        const element = {selector};
                        if (element) {{
                            element.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            element.click();
                            return true;
                        }}
                        return false;
                    }}''')
                    if button_clicked: return True

                # Handle XPath
                elif selector.startswith('//') or selector.startswith('(*'):
                    xpath_elements = page.locator(f"xpath={selector}")
                    if xpath_elements.count() > 0:
                        xpath_elements.first.click(force=True)
                        return True
                
                # Handle Standard CSS (Playwright auto-pierces Shadow DOM)
                else:
                    css_elements = page.locator(selector)
                    if css_elements.count() > 0:
                        css_elements.first.click(force=True)
                        return True

            except Exception:
                continue

        # Fallback JavaScript approach specifically for Element UI / Shadow DOM
        fallback_clicked = page.evaluate('''() => {
            // Helper to find element inside shadow roots recursively
            const findInShadow = (selector) => {
                let result = null;
                const search = (root) => {
                    if (root.querySelector(selector)) {
                        result = root.querySelector(selector);
                        return;
                    }
                    const shadows = Array.from(root.querySelectorAll('*')).filter(el => el.shadowRoot);
                    for (let s of shadows) {
                        search(s.shadowRoot);
                        if (result) return;
                    }
                };
                search(document);
                return result;
            };

            const calendarIcon = findInShadow('i.el-range__icon') || findInShadow('.el-icon-date');
            if (calendarIcon) {
                calendarIcon.scrollIntoView({behavior: 'auto', block: 'center'});
                calendarIcon.click();
                return true;
            }
            return false;
        }''')

        return fallback_clicked

    except Exception as e:
        print(f"Error in click_on_calendar: {str(e)}")
        return False

def click_on_yesterday(page):
    """
    Calculates yesterday's date and clicks it twice in the Element calendar.
    Handles Shadow DOM and dynamic date selection.
    """
    try:
        # 1. Calculate Yesterday's Day Number
        yesterday = datetime.now() - timedelta(days=1)
        day_to_click = str(yesterday.day)
        
        # 2. JavaScript to find and click the specific day inside Shadow DOM
        # We search for cells that are 'available' (not from prev/next month)
        # and contain the specific text of yesterday's day.
        click_script = f'''() => {{
            const findInShadow = (selector, root = document) => {{
                const el = root.querySelector(selector);
                if (el) return el;
                const shadows = Array.from(root.querySelectorAll('*')).filter(e => e.shadowRoot);
                for (let s of shadows) {{
                    const result = findInShadow(selector, s.shadowRoot);
                    if (result) return result;
                }}
                return null;
            }};

            // Find all available date cells
            const cells = Array.from(document.querySelectorAll('.el-date-table td.available'))
                          .concat(Array.from(findInShadow('.el-date-table') ? 
                                  findInShadow('.el-date-table').querySelectorAll('td.available') : []));

            // Filter for the cell that matches yesterday's date
            const targetCell = cells.find(cell => {{
                const text = cell.innerText.trim();
                return text === "{day_to_click}";
            }});

            if (targetCell) {{
                targetCell.scrollIntoView({{behavior: 'auto', block: 'center'}});
                // Click twice for range selection (Start and End)
                targetCell.click();
                setTimeout(() => targetCell.click(), 200); 
                return true;
            }}
            return false;
        }}'''

        success = page.evaluate(click_script)
        return success

    except Exception as e:
        print(f"Error in click_on_yesterday: {str(e)}")
        return False

def click_on_extrato_de_venda_next_page_button(page):
    """
    Finds and clicks the 'Next Page' button in the sales statement pagination.
    Handles Shadow DOM, checks for disabled states, and waits for UI transition.
    """
    try:
        # 1. Advanced JavaScript Evaluation for Shadow DOM and State
        # This is the most reliable method for Element UI components inside Shadow Roots
        js_click_script = '''() => {
            const findElementInShadows = (selector, root = document) => {
                const el = root.querySelector(selector);
                if (el) return el;
                const shadows = Array.from(root.querySelectorAll('*')).filter(e => e.shadowRoot);
                for (let s of shadows) {
                    const found = findElementInShadows(selector, s.shadowRoot);
                    if (found) return found;
                }
                return null;
            };

            const btn = findElementInShadows('button.btn-next');
            
            if (!btn) return "not_found";
            
            // Check if button is disabled via attribute, property, or CSS class
            const isReadonly = btn.disabled || 
                               btn.getAttribute('aria-disabled') === 'true' || 
                               btn.classList.contains('disabled');
            
            if (isReadonly) return "disabled";

            btn.scrollIntoView({behavior: 'auto', block: 'center'});
            btn.click();
            return "clicked";
        }'''

        result = page.evaluate(js_click_script)

        if result == "clicked":
            # Mandatory wait for the table to begin refreshing
            page.wait_for_timeout(2000) 
            return True
        elif result == "disabled":
            print("Pagination: Reached the last page (Next button is disabled).")
            return False
        else:
            # Fallback to Playwright's native locators if JS didn't find it
            # (Playwright pierces Shadow DOM automatically with CSS selectors)
            native_btn = page.locator("#pane-statement button.btn-next").first
            if native_btn.is_visible() and native_btn.is_enabled():
                native_btn.click(force=True)
                page.wait_for_timeout(2000)
                return True
            
            print("Pagination: Next page button not found.")
            return False

    except Exception as e:
        print(f"Error in click_on_extrato_de_venda_next_page_button: {str(e)}")
        return False

def click_on_gerar_relatorio_button(page):
    """
    Finds and clicks the 'Gerar Relatório' button in the sales statement.
    Handles Shadow DOM and checks for disabled state.
    """
    try:
        # List of selectors specific to the Gerar Relatório button
        selectors = [
            # 1. Direct CSS (Playwright automatically pierces Shadow DOM for CSS)
            "#pane-statement button.btn-primary:has-text('Gerar Relatório')",
            
            # 2. Your provided CSS Selector
            "#pane-statement > div > div:nth-child(1) > div > div.card-buttons > button",
            
            # 3. Text-based locator
            "button:has-text('Gerar Relatório')",
            
            # 4. Provided XPath
            "xpath=//*[@id='pane-statement']/div/div[1]/div/div[3]/button",
            
            # 5. Provided JS Path for Shadow DOM
            'document.querySelector("#privacy-web-myprivacy").shadowRoot.querySelector("#pane-statement > div > div:nth-child(1) > div > div.card-buttons > button")'
        ]

        for selector in selectors:
            try:
                # Handle JS Path specifically
                if selector.startswith("document.querySelector"):
                    clicked = page.evaluate(f'''() => {{
                        const btn = {selector};
                        if (btn && btn.getAttribute('aria-disabled') !== 'true') {{
                            btn.scrollIntoView({{behavior: 'smooth', block: 'center'}});
                            btn.click();
                            return true;
                        }}
                        return false;
                    }}''')
                    if clicked: return True

                # Handle Standard Locators (CSS/XPath)
                else:
                    loc = page.locator(selector).first
                    if loc.count() > 0 and loc.is_visible():
                        # Force visibility and click
                        loc.scroll_into_view_if_needed()
                        loc.click(force=True)
                        return True
            except:
                continue

        # Global Shadow DOM Fallback (Search for button by text inside all shadow roots)
        fallback_clicked = page.evaluate('''() => {
            const findInShadow = (root = document) => {
                const buttons = Array.from(root.querySelectorAll('button'));
                const target = buttons.find(b => b.innerText.includes('Gerar Relatório'));
                if (target && target.getAttribute('aria-disabled') !== 'true') return target;
                
                const shadows = Array.from(root.querySelectorAll('*')).filter(e => e.shadowRoot);
                for (let s of shadows) {
                    const found = findInShadow(s.shadowRoot);
                    if (found) return found;
                }
                return null;
            };
            const btn = findInShadow();
            if (btn) {
                btn.click();
                return true;
            }
            return false;
        }''')

        return fallback_clicked

    except Exception as e:
        print(f"Error in click_on_gerar_relatorio_button: {str(e)}")
        return False

def click_on_confirmar_button(page):
    """
    Finds and clicks the 'Confirmar' button inside the 'Gerar Relatório' overlay dialog.
    Handles the el-overlay-dialog structure specifically.
    """
    try:
        # 1. First, try to locate the specific dialog by its aria-label
        # This is the most reliable way to find this specific popup
        js_click_script = '''() => {
            const findButtonInDialog = (root = document) => {
                // 1. Find the dialog container by aria-label
                const dialog = root.querySelector('div[role="dialog"][aria-label="Gerar Relatório"]');
                
                if (dialog) {
                    // 2. Find the Confirmar button within that dialog's footer
                    const buttons = Array.from(dialog.querySelectorAll('button.btn-primary'));
                    const confirmBtn = buttons.find(b => b.innerText.includes('Confirmar'));
                    
                    if (confirmBtn && confirmBtn.getAttribute('aria-disabled') !== 'true') {
                        confirmBtn.click();
                        return "clicked";
                    }
                    return confirmBtn ? "disabled" : "button_not_found";
                }

                // 2. If not found, recurse into shadow roots
                const shadows = Array.from(root.querySelectorAll('*')).filter(e => e.shadowRoot);
                for (let s of shadows) {
                    const result = findButtonInDialog(s.shadowRoot);
                    if (result !== "not_found") return result;
                }
                return "not_found";
            };

            return findButtonInDialog();
        }'''

        result = page.evaluate(js_click_script)

        if result == "clicked":
            return True
        elif result == "disabled":
            print("Confirmar button is disabled (check if form fields are filled).")
            return False
        
        # 3. Native Playwright Fallback (If JS fails to find the dialog)
        # Playwright's locator('role=dialog') is very powerful
        try:
            confirm_loc = page.get_by_role("dialog", name="Gerar Relatório").get_by_role("button", name="Confirmar")
            if confirm_loc.is_visible():
                confirm_loc.click(force=True)
                return True
        except:
            pass

        print("Could not find the 'Gerar Relatório' confirmation dialog.")
        return False

    except Exception as e:
        print(f"Error in click_on_confirmar_button: {str(e)}")
        return False

def generate_top_spenders_from_report():
    """
    Reads the most recent XLSX report from the target folder, filters rows where Column E contains "Chat" or "Mimo - Chat",
    groups and sums Column D (Sua comissão) by unique values in Column H (Comprador), sorts by descending sum,
    saves the results to a new Excel file with yesterday's date in the format dd_mm_yyyy_top_spenders_privacy_vip.xlsx,
    and deletes the original report file.
    """
    target_folder = r"G:\Meu Drive\Sexting - Histórico"

    try:
        # 1. Find the most recent .xlsx file
        list_of_files = glob.glob(os.path.join(target_folder, "*.xlsx"))
        if not list_of_files:
            print("No XLSX files found in the directory.")
            return False

        latest_file = max(list_of_files, key=os.path.getctime)
        print(f"Reading latest Excel report: {os.path.basename(latest_file)}")

        # 2. Load the Workbook
        wb = load_workbook(latest_file, data_only=True)
        sheet = wb.active

        # Dictionary to hold sums: {Comprador: Valor_gasto}
        spenders_dict = {}

        # 3. Iterate through rows starting from row 2
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 8:  # Ensure row has at least columns A to H (indices 0-7)
                tipo_entrada = row[4]  # Column E (index 4)
                comprador = row[7]     # Column H (index 7)
                commission_raw = row[3]  # Column D (index 3)

                # Filter by "Chat" or "Mimo - Chat" in Column E
                if isinstance(tipo_entrada, str) and ("Chat" in tipo_entrada or "Mimo - Chat" in tipo_entrada):
                    if comprador is None or commission_raw is None:
                        continue

                    # Clean and parse the commission value
                    try:
                        if isinstance(commission_raw, str):
                            clean_val = commission_raw.replace('R$', '').replace('$', '').strip()
                            if ',' in clean_val and '.' in clean_val:
                                clean_val = clean_val.replace('.', '').replace(',', '.')
                            elif ',' in clean_val:
                                clean_val = clean_val.replace(',', '.')
                            commission = float(clean_val)
                        else:
                            commission = float(commission_raw)

                        # Group and sum by comprador
                        if comprador in spenders_dict:
                            spenders_dict[comprador] += commission
                        else:
                            spenders_dict[comprador] = commission

                    except (ValueError, TypeError):
                        continue

        # 4. Sort the dictionary by total commission descending
        sorted_spenders = sorted(spenders_dict.items(), key=lambda x: x[1], reverse=True)

        # 5. Create a new Workbook for output
        new_wb = Workbook()
        new_sheet = new_wb.active
        new_sheet.title = "Top Spenders"

        # Add headers
        new_sheet.append(["Comprador", "Total Commission"])

        # Add sorted data
        for comprador, total in sorted_spenders:
            new_sheet.append([comprador, total])

        # 6. Generate filename with yesterday's date
        yesterday = datetime.now() - timedelta(days=1)
        date_str = yesterday.strftime("%d_%m_%Y")
        output_filename = f"{date_str}_top_spenders_privacy_vip.xlsx"
        output_path = os.path.join(target_folder, output_filename)

        # Save the new workbook
        new_wb.save(output_path)
        print(f"Top spenders file saved to: {output_path}")

        # 7. CLOSE the original workbook before deleting
        wb.close()

        # 8. Delete the original file
        os.remove(latest_file)
        print(f"Original file {os.path.basename(latest_file)} deleted successfully.")

        return True

    except Exception as e:
        print(f"Error in generate_top_spenders_from_report: {str(e)}")
        return False

privacy_vip_income = 0.0

def main():
    pw = None
    context = None
    browser_process = None

    try:
        pw, context, page, browser_process = launch_chrome_with_debugging()

        print("Aguardando carregamento inicial da página...")
        page.wait_for_load_state("networkidle", timeout=60000)

        # region Closing pop-ups
        print("Verificando/fechando pop-ups...")
        for _ in range(5):
            if try_close_popup(page):
                time.sleep(1.2)
            else:
                break
        # endregion

        # region click on extrato mode
        print("\nTentando abrir aba Extrato...")
        extrato_success = False
        for attempt in range(1, 7):
            print(f"Tentativa {attempt}/6...")
            if click_extrato_tab(page):
                extrato_success = True
                break
            time.sleep(2)

        if not extrato_success:
            print("!!! FALHA CRÍTICA: Não conseguiu abrir a aba Extrato")
            return

        print("\nAguardando carregamento completo da tabela de extrato...")
        page.wait_for_selector('#pane-statement', state='visible', timeout=30000)
        time.sleep(3)  # segurança extra
        # endregion

        # region Try to click the Calendar icon
        max_retries = 3
        calendar_success = False

        for attempt in range(max_retries):
            if click_on_calendar(page):
                calendar_success = True
                break
            else:
                print(f"Calendar click attempt {attempt + 1} failed.")
                if attempt < max_retries - 1:
                    time.sleep(1)

        if not calendar_success:
            print("Failed to click Calendar after all attempts.")
        # endregion

        # region Try to click Yesterday with retries
        max_retries = 3
        for attempt in range(max_retries):
            if click_on_yesterday(page):
                print("Successfully selected Yesterday (Range Start & End).")
                break
            else:
                print(f"Attempt {attempt + 1} to click Yesterday failed.")
                time.sleep(1)
        else:
            print("Failed to select Yesterday after all attempts.")

        # endregion

        time.sleep(4)

        # region Try to click the Gerar Relatório button
        if click_on_gerar_relatorio_button(page):
            print("Dialog 'Gerar Relatório' opened.")
            time.sleep(2) # Wait for dialog animation
            
            # DO NOT call click_on_confirmar_button here!
            # It will be called inside the download listener below.
        # endregion

                # region Confirm and Download Report
        target_folder = r"G:\Meu Drive\Sexting - Histórico"  # Adjusted to match the processing folder
        os.makedirs(target_folder, exist_ok=True)
        max_retries = 3
        download_success = False
        for attempt in range(max_retries):
            try:
                print(f"Attempt {attempt + 1}: Waiting for download event...")
                # 1. Start the listener
                with page.expect_download(timeout=60000) as download_info:
                    # 2. Trigger the click ONLY HERE
                    if click_on_confirmar_button(page):
                        print("Confirm button clicked. Processing file...")
                        # 3. Capture and save the file
                        download = download_info.value
                        final_destination = os.path.join(target_folder, download.suggested_filename)
                        download.save_as(final_destination)
                        print(f"File successfully saved to: {final_destination}")
                        download_success = True
                        break 
                    else:
                        print(f"Click logic could not find the button on attempt {attempt + 1}")
            except Exception as e:
                print(f"Attempt {attempt + 1} failed: {str(e)}")
                if attempt < max_retries - 1:
                    time.sleep(2)
        if not download_success:
            print("Failed to capture the download after all retries.")
        # endregion

        # After successful download, generate top spenders
        if download_success:
            generate_top_spenders_from_report()

        print("\nTask completed successfully. Closing automation...")

    except Exception as e:
        print(f"\nEXECUTION ERROR:\n{type(e).__name__}: {e}")
    
    finally:
        # This block runs no matter what, ensuring no zombie chrome processes remain
        print("Cleaning up resources...")
        try:
            context.close()
            pw.stop()
            # Terminate the background process
            browser_process.terminate() 
        except:
            pass
        print("Program finished.")

if __name__ == "__main__":
    main()